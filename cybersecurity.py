import datetime
from datetime import date
import sqlite3
from openpyxl import Workbook, load_workbook
import csv
import os
import shutil
import pandas as pd

def readfilename():
    #this is function to list file name of files in director. The function will return a list of names
    filename_csv=[]
    for x in os.listdir():
        if x.endswith('.csv'):
            filename_csv.append(x)
    return filename_csv

def movefile(filename):
    currenttime = datetime.datetime.now()
    newname=str(currenttime)[:10]+'_'+str(currenttime)[-2:]+"_"+filename
    newdir='archive/'+newname
    os.rename(filename,newdir)

def removedup():
    #this will remove duplicated row by importing one file twice
    conn = sqlite3.connect('secure_database.db')
    c = conn.cursor()
    c.execute('DELETE FROM security WHERE rowid NOT IN (SELECT MIN(rowid) FROM security GROUP BY id,cve,cvss,risk,host,protocol,port,name,synopsis,description,solution,seealso,output,filename)')
    conn.commit()
    conn.close()
    
def buildfile():
    #you can import same data mutiple times by accident, but one the oldest one will be saved in the database.
    conn=sqlite3.connect('secure_database.db')
    c=conn.cursor()
    #Creat the table named security
    c.execute('CREATE table IF NOT EXISTS security (id INTEGER, cve TEXT,cvss TEXT,risk TEXT,host TEXT,protocol TEXT,port INTEGER,name TEXT,synopsis TEXT,description TEXT,solution TEXT,seealso TEXT,output TEXT,solved TEXT NOT NULL DEFAULT "No" ,comment TEXT,filename TEXT,timestamp TEXT)')
    #read csv files
    filenames=readfilename()
    for file_csv in filenames:
        df=pd.read_csv(file_csv)


        #insert the csv files to table
        for index, row in df.iterrows():
            insert_records='INSERT INTO security(id,cve,cvss,risk,host,protocol,port,name,synopsis,description,solution,seealso,output) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)'
            c.execute(insert_records,(row['Plugin ID'],row['CVE'],row['CVSS v2.0 Base Score'],row['Risk'],row['Host'],row['Protocol'],row['Port'],row['Name'],row['Synopsis'],row['Description'],row['Solution'],row['See Also'],row['Plugin Output']))
            #update the filename and timestamp column
            #first get the time.
            currenttime=datetime.datetime.now()
            updatecsv='UPDATE security SET filename=UPPER("{fname}"),timestamp="{tstamp}" WHERE filename IS NULL'.format(fname=file_csv,tstamp=currenttime)
            c.execute(updatecsv)
        conn.commit()

    conn.close()
    #move csv file to archive folder after importing data.
    #if movefile failed, then just pass it.
    for file_csv in filenames:
        try:
            movefile(file_csv)
        except:
            pass
    
    #remove duplication
    removedup()
    

def update_solved(id,value):
    #this will remove duplicated row by importing one file twice
    conn = sqlite3.connect('secure_database.db')
    c = conn.cursor()

    update_solved="UPDATE security SET solved='{value}' WHERE rowid='{id}'".format(value=value,id=id)

    c.execute(update_solved)
    conn.commit()
    conn.close()

def update_comment(id,value):
    #this will remove duplicated row by importing one file twice
    conn = sqlite3.connect('secure_database.db')
    c = conn.cursor()
    update_solved="UPDATE security SET comment='{value}' WHERE rowid='{id}'".format(value=value,id=id)
    print (update_solved)
    c.execute(update_solved)
    conn.commit()
    conn.close()


def downloadfile():
    result_data=dataread()
    wb=load_workbook(r"")
    ws=wb.active
    ws.cell(row=1, column=1).value = "P-number"
    ws.cell(row=1, column=2).value = "Branch"
    ws.cell(row=1, column=3).value = "Wrong Cost"
    ws.cell(row=1, column=4).value = "Wrong Package"
    ws.cell(row=1, column=5).value = "Wrong Part"
    ws.cell(row=1, column=6).value = "Wrong Under M/N"
    ws.cell(row=1, column=7).value = "Other"
    ws.cell(row=1, column=8).value = "Additon"
    ws.cell(row=1, column=9).value = "Date"

    i=1
    for s in result_data:
        i=i+1
        ws.cell(row=i,column=1).value=s[0]
        ws.cell(row=i, column=2).value = s[1]
        ws.cell(row=i, column=3).value = s[2]
        ws.cell(row=i, column=4).value = s[3]
        ws.cell(row=i, column=5).value = s[4]
        ws.cell(row=i, column=6).value = s[5]
        ws.cell(row=i, column=7).value = s[6]
        ws.cell(row=i, column=8).value = s[7]
        ws.cell(row=i, column=9).value = s[8]
    wb.save("myfile\wrong_purchase_trace.xlsx")





def know(rowid):
    #this will remove duplicated row by importing one file twice
    conn = sqlite3.connect('secure_database.db')
    c = conn.cursor()
    search_name = "SELECT name FROM security WHERE rowid='{rowid}'".format(rowid=rowid)
    search_condition="SELECT rowid,* FROM security WHERE name=({name}) AND solved='Yes' ORDER BY timestamp DESC ".format(name=search_name)
    # print (search_condition)
    c.execute(search_condition)
    results = c.fetchall()
    conn.commit()
    conn.close()
    return results


